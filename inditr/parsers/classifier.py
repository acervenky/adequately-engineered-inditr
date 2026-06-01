"""
LLM-based document classifier.

Replaces the N-parser can_parse() chain with a single LLM call per uploaded file.

Flow:
    filepath
        → extract text preview (filename + first ~3000 chars of content)
        → LLM: "what type of document is this?"
        → ClassificationResult { doc_type, confidence, reason }
        → registry routes to the right parser directly

Why this beats per-parser regex sniffers:
    - One LLM call instead of N regex scans
    - Filename alone is often enough ("taxpnl-VC5670-2025_2026.xlsx" → Zerodha)
    - First 3000 chars of content is definitive for all formats we've seen
    - Identifies document types we have no parser for yet (routes to vision fallback
      but at least tells the user what they uploaded and what's missing)
    - Handles Zerodha CSV, Upstox F&O, CAMS statements, Form 26AS — all without
      writing format-specific sniffers

Confidence threshold: if LLM confidence < 0.70, fall through to legacy can_parse()
chain as a safety net. Never fail completely.
"""
from __future__ import annotations
import json
import logging
import os
import re
from dataclasses import dataclass

logger = logging.getLogger(__name__)

# LLM confidence below this → fall through to legacy can_parse() chain
_MIN_CONFIDENCE = 0.70

# Known document types the LLM should choose from.
# Keeping this list explicit so the LLM doesn't invent types.
_KNOWN_TYPES = [
    "zerodha_tax_pnl",        # Zerodha Tax P&L (xlsx multi-sheet or CSV)
    "upstox_realized_pnl",    # Upstox Realized P&L (xlsx or PDF)
    "form_16",                # Employer TDS certificate Part A + Part B (PDF)
    "form_26as",              # Annual TDS/TCS statement from IT portal (PDF/text)
    "salary_slip",            # Monthly payslip (PDF)
    "bank_statement",         # Bank account statement (PDF)
    "cams_mf_statement",      # CAMS/KFintech consolidated MF statement (PDF)
    "nps_statement",          # NPS PRAN statement (PDF)
    "home_loan_certificate",  # Interest/principal certificate from lender (PDF)
    "unknown",                # Cannot determine — route to vision fallback
]

_TYPE_DESCRIPTIONS = {
    "zerodha_tax_pnl":       "Zerodha Tax P&L report; multi-sheet xlsx with Equity/F&O/Tradewise sheets, or CSV with ISIN/STCG/LTCG columns; filename often 'taxpnl-CLIENTID-YYYY_YYYY.xlsx'",
    "upstox_realized_pnl":   "Upstox Realized P&L report; xlsx with sheet REALIZED_PNL_DOWNLOAD or PDF; columns include Buy Date, Sell Date, Short Term, Long Term, Speculation; filename often 'realizedPnL_EQ_...'",
    "form_16":               "Form 16 TDS certificate issued by employer; PDF; contains TRACES header, employer TAN, Part A quarterly TDS table, Part B salary breakup and Chapter VI-A deductions",
    "form_26as":             "Annual Information Statement or Form 26AS from Income Tax portal; PDF or text; contains TDS from all deductors, advance tax, SFT transactions",
    "salary_slip":           "Monthly salary payslip from employer; PDF; shows earnings/deductions breakdown for a single month, usually no TAN or TRACES header",
    "bank_statement":        "Bank account statement; PDF; shows credit/debit transactions with dates; bank name prominent (HDFC/SBI/ICICI/Axis/Kotak etc.)",
    "cams_mf_statement":     "CAMS or KFintech consolidated mutual fund account statement; PDF; lists folios, NAV, units across multiple AMCs",
    "nps_statement":         "NPS PRAN statement from NSDL/CAMS; PDF; shows PRAN number, contributions, NAV-based balance",
    "home_loan_certificate": "Home loan interest/principal certificate from bank or NBFC; PDF; shows loan account, principal repaid, interest paid for the year",
    "unknown":               "Document type cannot be determined from filename and content preview",
}


@dataclass
class ClassificationResult:
    doc_type: str           # one of _KNOWN_TYPES
    confidence: float       # 0.0–1.0
    reason: str             # brief LLM explanation
    is_llm: bool = True     # False if fell back to legacy can_parse()


# ── Preview extraction ────────────────────────────────────────────────────────

def _extract_preview(filepath: str, max_chars: int = 3000) -> str:
    """
    Extract a text preview from the file.
    PDF → pdfplumber first 2 pages.
    XLSX → openpyxl sheet names + first 20 rows of first sheet.
    CSV/TXT → direct read.
    Returns empty string on any error (classifier will rely on filename alone).
    """
    fp = filepath.lower()
    try:
        if fp.endswith(".pdf"):
            import pdfplumber
            lines: list[str] = []
            with pdfplumber.open(filepath) as pdf:
                for page in pdf.pages[:2]:
                    text = page.extract_text() or ""
                    lines.append(text)
                    if sum(len(l) for l in lines) >= max_chars:
                        break
            return "\n".join(lines)[:max_chars]

        elif fp.endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            parts = [f"Sheet names: {', '.join(wb.sheetnames)}"]
            # First sheet content
            ws = wb.worksheets[0]
            row_lines: list[str] = []
            for row in ws.iter_rows(max_row=25, values_only=True):
                cells = [str(c or "").strip() for c in row if c is not None]
                if cells:
                    row_lines.append(" | ".join(cells))
            parts.append("\n".join(row_lines))
            wb.close()
            return "\n".join(parts)[:max_chars]

        elif fp.endswith((".csv", ".txt")):
            with open(filepath, "r", encoding="utf-8", errors="replace") as f:
                return f.read(max_chars)

    except Exception as exc:
        logger.debug("Preview extraction failed for %s: %s", filepath, exc)

    return ""


# ── LLM classification ────────────────────────────────────────────────────────

def classify_document(filepath: str) -> ClassificationResult:
    """
    Classify a document using filename + content preview via LLM.

    Returns ClassificationResult. Never raises — returns doc_type="unknown"
    with confidence=0.0 on any failure.
    """
    filename = os.path.basename(filepath)
    preview = _extract_preview(filepath)

    type_list = "\n".join(
        f'  "{k}": {v}' for k, v in _TYPE_DESCRIPTIONS.items()
    )

    prompt = f"""You are classifying an Indian income tax document.

Filename: {filename}

Content preview (first ~3000 chars):
---
{preview or "(could not extract text preview)"}
---

Classify this document as exactly one of these types:
{type_list}

Rules:
- Choose the MOST SPECIFIC matching type
- Use "unknown" only if genuinely ambiguous even with the content preview
- The filename is often the strongest signal (e.g. "taxpnl" → zerodha_tax_pnl)
- If the content shows Zerodha sheet names (Equity, F&O, Tradewise), confidence should be 0.98+

Respond with JSON only:
{{"doc_type": "<type>", "confidence": <0.0-1.0>, "reason": "<one sentence>"}}"""

    try:
        import litellm
        from inditr.graph.llm import MODEL

        response = litellm.completion(
            model=MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.0,
            max_tokens=150,
        )
        raw = response.choices[0].message.content or ""
        cleaned = re.sub(r"```(?:json)?\s*", "", raw).strip().rstrip("`").strip()
        data = json.loads(cleaned)

        doc_type = data.get("doc_type", "unknown")
        if doc_type not in _KNOWN_TYPES:
            doc_type = "unknown"

        return ClassificationResult(
            doc_type=doc_type,
            confidence=float(data.get("confidence", 0.0)),
            reason=str(data.get("reason", "")),
            is_llm=True,
        )

    except Exception as exc:
        logger.warning("LLM classification failed for %s: %s", filename, exc)
        return ClassificationResult(
            doc_type="unknown",
            confidence=0.0,
            reason=f"LLM classification failed: {exc}",
            is_llm=False,
        )
