"""
Upstox Realized P&L parser.

Supported formats:
  1. XLSX — "realizedPnL_EQ_YYYY-MM-DD_To_YYYY-MM-DD_CLIENTID.xlsx"
     Single sheet "REALIZED_PNL_DOWNLOAD".
     Header structure:
       Rows 1-8  : company info, UCC, Name, PAN, report period
       Rows 10-12: summary (Gross P&L, Net P&L)
       Rows 14-24: charges breakdown
       Row 26    : trade table header
       Row 27+   : per-trade rows
     Columns: Scrip Name, Scrip Code, Symbol, ISIN, Scrip Opt, Qty,
              Buy Date, Buy Rate, Buy Amt, Sell Date, Sell Rate, Sell Amt,
              Days, Total PL, Short Term, Long Term, Speculation, Turn Over

  2. PDF — same report exported as PDF (pdfplumber tables)
     Same columns as xlsx; table spans multiple pages.

Classification:
  Speculation column non-null → intraday/speculation
  Short Term column non-null  → STCG (Days < 365 confirmed from Days col)
  Long Term column non-null   → LTCG (Days >= 365)

Parsers NEVER raise unhandled exceptions.
"""
from __future__ import annotations
import re
import traceback
from typing import Optional

import pdfplumber

from inditr.models.documents import ExtractedField, ParsedDocument
from inditr.models.tax_data import GainType, AssetType
from .base import BaseParser

_EXACT_CONF = 0.95
_FUZZY_CONF = 0.75

# Actual column headers in Upstox reports (strip + lower for matching)
_TRADE_HEADERS = [
    "scrip name", "buy date", "buy amt", "sell date", "sell amt",
    "days", "total pl", "short term", "long term", "speculation",
]


def _to_float(val) -> float:
    """Convert cell value (numeric or string) to float. Returns 0.0 on failure."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    val = str(val).replace(",", "").replace("(", "-").replace(")", "").strip()
    try:
        return float(val)
    except ValueError:
        return 0.0


class UpstoxParser(BaseParser):

    def can_parse(self, filepath: str) -> bool:
        fp = filepath.lower()
        if fp.endswith(".xlsx"):
            return self._sniff_xlsx(filepath)
        if fp.endswith(".pdf"):
            return self._sniff_pdf(filepath)
        return False

    def parse(self, filepath: str) -> ParsedDocument:
        doc = ParsedDocument(doc_type="upstox_pnl", filename=filepath)
        try:
            if filepath.lower().endswith(".xlsx"):
                self._parse_xlsx(filepath, doc)
            else:
                self._parse_pdf(filepath, doc)
        except Exception as exc:
            doc.parse_errors.append(
                f"Unhandled error in UpstoxParser: {exc}\n{traceback.format_exc()}"
            )
        return doc

    # ── Sniffers ──────────────────────────────────────────────────────────────

    def _sniff_xlsx(self, filepath: str) -> bool:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            has_sheet = "REALIZED_PNL_DOWNLOAD" in wb.sheetnames
            wb.close()
            if has_sheet:
                return True
            # Fallback: check filename pattern
            import os
            return "realizedpnl" in os.path.basename(filepath).lower() or "upstox" in filepath.lower()
        except Exception:
            return False

    def _sniff_pdf(self, filepath: str) -> bool:
        try:
            with pdfplumber.open(filepath) as pdf:
                text = ""
                for page in pdf.pages[:2]:
                    text += (page.extract_text() or "").lower()
            return "upstox" in text or (
                "gross p&l" in text and ("buy date" in text or "sell date" in text)
            )
        except Exception:
            return False

    # ── XLSX — primary format ─────────────────────────────────────────────────

    def _parse_xlsx(self, filepath: str, doc: ParsedDocument) -> None:
        import openpyxl
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as exc:
            doc.parse_errors.append(f"Cannot open XLSX: {exc}")
            return

        sheet_name = "REALIZED_PNL_DOWNLOAD"
        if sheet_name not in wb.sheetnames:
            # Try first sheet
            sheet_name = wb.sheetnames[0]
            doc.parse_errors.append(
                f"Expected sheet 'REALIZED_PNL_DOWNLOAD' not found; using '{sheet_name}'"
            )

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # ── Header info (scan col A for key-value pairs) ───────────────────
        for row in rows[:10]:
            label = str(row[0] or "").strip().lower()
            value = row[1] if len(row) > 1 else None
            if label == "pan" and value:
                doc.fields["pan"] = ExtractedField(
                    value=str(value).strip(), source_document=doc.filename,
                    source_section="xlsx:header", confidence=_EXACT_CONF,
                )
            elif label == "ucc" and value:
                doc.fields["client_id"] = ExtractedField(
                    value=str(value).strip(), source_document=doc.filename,
                    source_section="xlsx:header", confidence=_EXACT_CONF,
                )
            elif label == "name" and value:
                doc.fields["client_name"] = ExtractedField(
                    value=str(value).strip(), source_document=doc.filename,
                    source_section="xlsx:header", confidence=_EXACT_CONF,
                )

        # ── Summary (Gross P&L, Net P&L) ──────────────────────────────────
        for row in rows:
            label = str(row[0] or "").strip().lower()
            value = row[1] if len(row) > 1 else None
            if label == "gross p&l" and value is not None:
                doc.fields["gross_pnl"] = ExtractedField(
                    value=_to_float(value), source_document=doc.filename,
                    source_section="xlsx:summary", confidence=_EXACT_CONF,
                )
            elif label == "net p&l" and value is not None:
                doc.fields["net_pnl"] = ExtractedField(
                    value=_to_float(value), source_document=doc.filename,
                    source_section="xlsx:summary", confidence=_EXACT_CONF,
                )

        # ── Trade table ────────────────────────────────────────────────────
        hdr_idx = self._find_trade_header(rows)
        if hdr_idx is None:
            doc.parse_errors.append("Upstox XLSX: could not find trade table header row")
            return

        headers = [str(c or "").strip().lower() for c in rows[hdr_idx]]
        gains, stcg_total, ltcg_total, speculation_total = self._parse_trade_rows(
            rows[hdr_idx + 1:], headers, source="xlsx",
        )

        self._set_fields(doc, gains, stcg_total, ltcg_total, speculation_total, source_section="xlsx")

    # ── PDF — secondary format ────────────────────────────────────────────────

    def _parse_pdf(self, filepath: str, doc: ParsedDocument) -> None:
        try:
            pdf = pdfplumber.open(filepath)
        except Exception as exc:
            doc.parse_errors.append(f"Cannot open PDF: {exc}")
            return

        all_rows: list[list] = []
        with pdf:
            for page in pdf.pages:
                try:
                    for table in (page.extract_tables() or []):
                        all_rows.extend(table or [])
                except Exception as exc:
                    doc.parse_errors.append(f"Page extraction error: {exc}")

        # Client info from text (PAN pattern)
        with pdfplumber.open(filepath) as pdf:
            full_text = "\n".join(
                page.extract_text() or "" for page in pdf.pages[:2]
            )
        pan_m = re.search(r"\bPAN\s*[:\-]?\s*([A-Z]{5}[0-9]{4}[A-Z])\b", full_text)
        if pan_m:
            doc.fields["pan"] = ExtractedField(
                value=pan_m.group(1), source_document=doc.filename,
                source_section="pdf:text", confidence=_EXACT_CONF,
            )

        hdr_idx = self._find_trade_header(all_rows)
        if hdr_idx is None:
            doc.parse_errors.append("Upstox PDF: could not find trade table header row")
            return

        headers = [str(c or "").strip().lower() for c in all_rows[hdr_idx]]
        # PDF sometimes splits "Scrip\nOpt" into two lines in the cell
        headers = [h.replace("\n", " ") for h in headers]

        gains, stcg_total, ltcg_total, speculation_total = self._parse_trade_rows(
            all_rows[hdr_idx + 1:], headers, source="pdf",
        )

        self._set_fields(doc, gains, stcg_total, ltcg_total, speculation_total, source_section="pdf")

    # ── Shared helpers ────────────────────────────────────────────────────────

    def _find_trade_header(self, rows: list) -> Optional[int]:
        """Find the row index of the trade table header."""
        for i, row in enumerate(rows):
            cells = [str(c or "").strip().lower().replace("\n", " ") for c in row]
            matched = sum(1 for h in _TRADE_HEADERS if any(h in c for c in cells))
            if matched >= 4:
                return i
        return None

    def _parse_trade_rows(
        self,
        rows: list,
        headers: list[str],
        source: str,
    ) -> tuple[list[dict], float, float, float]:
        """Parse individual trade rows into gain records."""

        def _col(name: str) -> Optional[int]:
            for i, h in enumerate(headers):
                if name in h:
                    return i
            return None

        c_scrip       = _col("scrip name")
        c_isin        = _col("isin")
        c_buy_date    = _col("buy date")
        c_buy_amt     = _col("buy amt")
        c_sell_date   = _col("sell date")
        c_sell_amt    = _col("sell amt")
        c_days        = _col("days")
        c_short_term  = _col("short term")
        c_long_term   = _col("long term")
        c_speculation = _col("speculation")

        gains: list[dict] = []
        stcg_total = 0.0
        ltcg_total = 0.0
        speculation_total = 0.0

        for row in rows:
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            def _cell(col_idx):
                if col_idx is None or col_idx >= len(row):
                    return None
                return row[col_idx]

            scrip     = str(_cell(c_scrip) or "").strip()
            isin      = str(_cell(c_isin) or "").strip()
            buy_date  = str(_cell(c_buy_date) or "").strip()
            sell_date = str(_cell(c_sell_date) or "").strip()
            buy_amt   = _to_float(_cell(c_buy_amt))
            sell_amt  = _to_float(_cell(c_sell_amt))
            days      = _to_float(_cell(c_days))
            stcg_val  = _to_float(_cell(c_short_term))
            ltcg_val  = _to_float(_cell(c_long_term))
            spec_val  = _to_float(_cell(c_speculation))

            if not scrip and not isin:
                continue

            asset_type = AssetType.EQUITY  # Upstox EQ report is equity only

            if spec_val != 0:
                speculation_total += spec_val
                gains.append({
                    "gain_type": GainType.STCG,
                    "asset_type": asset_type,
                    "isin": isin,
                    "scrip": scrip,
                    "gain_amount": spec_val,
                    "buy_date": buy_date,
                    "sell_date": sell_date,
                    "buy_value": buy_amt,
                    "sell_value": sell_amt,
                    "holding_days": int(days),
                    "is_speculation": True,
                })

            if stcg_val != 0:
                stcg_total += stcg_val
                gains.append({
                    "gain_type": GainType.STCG,
                    "asset_type": asset_type,
                    "isin": isin,
                    "scrip": scrip,
                    "gain_amount": stcg_val,
                    "buy_date": buy_date,
                    "sell_date": sell_date,
                    "buy_value": buy_amt,
                    "sell_value": sell_amt,
                    "holding_days": int(days),
                    "is_speculation": False,
                })

            if ltcg_val != 0:
                ltcg_total += ltcg_val
                gains.append({
                    "gain_type": GainType.LTCG,
                    "asset_type": asset_type,
                    "isin": isin,
                    "scrip": scrip,
                    "gain_amount": ltcg_val,
                    "buy_date": buy_date,
                    "sell_date": sell_date,
                    "buy_value": buy_amt,
                    "sell_value": sell_amt,
                    "holding_days": int(days),
                    "is_speculation": False,
                })

        return gains, stcg_total, ltcg_total, speculation_total

    def _set_fields(
        self,
        doc: ParsedDocument,
        gains: list[dict],
        stcg_total: float,
        ltcg_total: float,
        speculation_total: float,
        source_section: str,
    ) -> None:
        doc.fields["stcg_equity_total"] = ExtractedField(
            value=stcg_total, source_document=doc.filename,
            source_section=source_section, confidence=_EXACT_CONF,
        )
        doc.fields["ltcg_equity_total"] = ExtractedField(
            value=ltcg_total, source_document=doc.filename,
            source_section=source_section, confidence=_EXACT_CONF,
        )
        doc.fields["speculation_total"] = ExtractedField(
            value=speculation_total, source_document=doc.filename,
            source_section=source_section, confidence=_EXACT_CONF,
        )
        doc.fields["capital_gains"] = ExtractedField(
            value=gains, source_document=doc.filename,
            source_section=source_section,
            confidence=_EXACT_CONF if gains else _FUZZY_CONF,
        )

        if not gains:
            doc.parse_errors.append("No capital gain rows extracted from Upstox P&L")
