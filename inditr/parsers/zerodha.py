"""
Zerodha Tax P&L parser.

Supported formats (in priority order):
  1. XLSX — Zerodha's "taxpnl-CLIENTID-YYYY_YYYY-Q1-Q4.xlsx"
     Primary format. Multi-sheet workbook:
       "Equity"                  → STCG / LTCG / speculation summary totals
       "F&O"                     → Options + Futures realized profit
       "Tradewise Exits from..."  → Per-trade rows; holding period classifies STCG vs LTCG
       "Mutual Funds"            → MF STCG/LTCG totals
  2. PDF  — pdfplumber table extraction (older export format)
  3. CSV  — direct CSV export (legacy)

All formats extract the same fields into ParsedDocument:
  stcg_equity_total, ltcg_equity_total, speculation_total,
  fno_total, stcg_mf_total, ltcg_mf_total,
  capital_gains (list of per-trade dicts), pan, client_id

Parsers NEVER raise unhandled exceptions.
"""
from __future__ import annotations
import csv
import io
import re
import traceback
from typing import Optional

import pdfplumber

from inditr.models.documents import ExtractedField, ParsedDocument
from inditr.models.tax_data import GainType, AssetType
from .base import BaseParser

_EXACT_CONF = 0.95
_FUZZY_CONF = 0.75

# Scheme-name keywords that unambiguously identify a debt mutual fund.
# All post-Apr-2023 debt MF gains are taxed at slab rates (Finance Act 2023).
# Equity MF keywords ("equity", "elss", "arbitrage", etc.) are NOT in this set
# so equity funds never accidentally match.
_DEBT_MF_KEYWORDS: frozenset[str] = frozenset({
    "liquid", "overnight", "money market",
    "ultra short", "low duration", "short duration", "medium duration",
    "long duration", "dynamic bond", "corporate bond", "credit risk",
    "banking and psu", "banking & psu", "gilt", "g-sec",
    "fixed maturity", "fmp", "floater", "floating rate",
    "debt fund", "debt index", "income fund", "treasury",
})


# ── Helpers ───────────────────────────────────────────────────────────────────

def _isin_to_asset_type(isin: Optional[str], scheme_name: str = "") -> AssetType:
    """
    Classify an MF/equity ISIN into AssetType.

    For INF/0P ISINs (mutual funds), uses the scheme name to distinguish equity
    MF (Section 112A, 12.5%/20%) from debt MF (slab rate, Finance Act 2023).
    Falls back to EQUITY_MF when scheme_name is empty or ambiguous — the
    gap_fill debt-MF question catches any remaining uncertain cases.
    """
    if not isin:
        return AssetType.OTHER
    isin = str(isin).strip().upper()
    if isin.startswith("INF") or isin.startswith("0P"):
        name_lower = scheme_name.lower()
        if any(kw in name_lower for kw in _DEBT_MF_KEYWORDS):
            return AssetType.DEBT_MF
        return AssetType.EQUITY_MF
    if isin.startswith("INE"):
        return AssetType.EQUITY
    return AssetType.OTHER


def _to_float(val) -> float:
    """Convert openpyxl cell value or string to float. Returns 0.0 on failure."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    val = str(val).replace(",", "").replace("(", "-").replace(")", "").strip()
    try:
        return float(val)
    except ValueError:
        return 0.0


def _find_row_by_label(rows: list, label: str, col: int = 1) -> Optional[tuple]:
    """Return first row whose col-th element matches label (case-insensitive)."""
    label_lower = label.lower().strip()
    for row in rows:
        cell = str(row[col] or "").strip().lower() if len(row) > col else ""
        if cell == label_lower:
            return row
    return None


def _find_header_row(rows: list, required: list[str], col_start: int = 1) -> Optional[int]:
    """Return index of the row that contains all required header labels."""
    req_lower = [r.lower() for r in required]
    for i, row in enumerate(rows):
        cells = [str(c or "").strip().lower() for c in row[col_start:]]
        if all(any(r in c for c in cells) for r in req_lower):
            return i
    return None


# ── Parser ────────────────────────────────────────────────────────────────────

class ZerodhaPnlParser(BaseParser):

    def can_parse(self, filepath: str) -> bool:
        fp = filepath.lower()
        if fp.endswith(".xlsx"):
            return self._sniff_xlsx(filepath)
        if fp.endswith(".csv"):
            return self._sniff_csv(filepath)
        if fp.endswith(".pdf"):
            return self._sniff_pdf(filepath)
        return False

    def parse(self, filepath: str) -> ParsedDocument:
        doc = ParsedDocument(doc_type="zerodha_pnl", filename=filepath)
        try:
            fp = filepath.lower()
            if fp.endswith(".xlsx"):
                self._parse_xlsx(filepath, doc)
            elif fp.endswith(".csv"):
                self._parse_csv(filepath, doc)
            else:
                self._parse_pdf(filepath, doc)
        except Exception as exc:
            doc.parse_errors.append(
                f"Unhandled error in ZerodhaPnlParser: {exc}\n{traceback.format_exc()}"
            )
        return doc

    # ── Sniffers ──────────────────────────────────────────────────────────────

    def _sniff_xlsx(self, filepath: str) -> bool:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            sheet_names = [s.lower() for s in wb.sheetnames]
            wb.close()
            return (
                any("equity" in s for s in sheet_names)
                and any("tradewise" in s or "f&o" in s for s in sheet_names)
            )
        except Exception:
            return False

    def _sniff_csv(self, filepath: str) -> bool:
        try:
            with open(filepath, "r", encoding="utf-8", errors="replace") as f:
                header = f.read(500).lower()
            return "zerodha" in header or (
                "isin" in header and ("stcg" in header or "ltcg" in header or "short term" in header)
            )
        except Exception:
            return False

    def _sniff_pdf(self, filepath: str) -> bool:
        try:
            with pdfplumber.open(filepath) as pdf:
                text = ""
                for page in pdf.pages[:2]:
                    text += (page.extract_text() or "").lower()
            return "zerodha" in text or (
                "profit & loss" in text and ("isin" in text or "stcg" in text)
            )
        except Exception:
            return False

    # ── XLSX — primary format ─────────────────────────────────────────────────

    def _parse_xlsx(self, filepath: str, doc: ParsedDocument) -> None:
        import openpyxl
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as exc:
            doc.parse_errors.append(f"Cannot open XLSX: {exc}")
            return

        # ── Client info (PAN, name) from any summary sheet ────────────────────
        for sheet_name in wb.sheetnames:
            rows = list(wb[sheet_name].iter_rows(values_only=True))
            pan_row = _find_row_by_label(rows, "PAN", col=1)
            if pan_row and len(pan_row) > 2 and pan_row[2]:
                doc.fields["pan"] = ExtractedField(
                    value=str(pan_row[2]).strip(),
                    source_document=doc.filename,
                    source_section=f"xlsx:{sheet_name}",
                    confidence=_EXACT_CONF,
                )
            cid_row = _find_row_by_label(rows, "Client ID", col=1)
            if cid_row and len(cid_row) > 2 and cid_row[2]:
                doc.fields["client_id"] = ExtractedField(
                    value=str(cid_row[2]).strip(),
                    source_document=doc.filename,
                    source_section=f"xlsx:{sheet_name}",
                    confidence=_EXACT_CONF,
                )
            if "pan" in doc.fields:
                break

        # ── Equity summary ────────────────────────────────────────────────────
        eq_sheet = next((wb[n] for n in wb.sheetnames if n.lower() == "equity"), None)
        speculation_total = 0.0
        stcg_equity_total = 0.0
        ltcg_equity_total = 0.0

        if eq_sheet:
            rows = list(eq_sheet.iter_rows(values_only=True))
            r = _find_row_by_label(rows, "Intraday/Speculative profit", col=1)
            if r:
                speculation_total = _to_float(r[2] if len(r) > 2 else None)
            r = _find_row_by_label(rows, "Short Term profit", col=1)
            if r:
                stcg_equity_total = _to_float(r[2] if len(r) > 2 else None)
            r = _find_row_by_label(rows, "Long Term profit", col=1)
            if r:
                ltcg_equity_total = _to_float(r[2] if len(r) > 2 else None)
        else:
            doc.parse_errors.append("Zerodha XLSX: 'Equity' sheet not found")

        # ── F&O summary ───────────────────────────────────────────────────────
        fno_sheet = next((wb[n] for n in wb.sheetnames if n.lower() == "f&o"), None)
        fno_total = 0.0

        if fno_sheet:
            rows = list(fno_sheet.iter_rows(values_only=True))
            r = _find_row_by_label(rows, "Options Realized Profit", col=1)
            if r:
                fno_total += _to_float(r[2] if len(r) > 2 else None)
            r = _find_row_by_label(rows, "Futures Realized Profit", col=1)
            if r:
                fno_total += _to_float(r[2] if len(r) > 2 else None)

        # ── Mutual Funds summary ──────────────────────────────────────────────
        mf_sheet = next((wb[n] for n in wb.sheetnames if n.lower() == "mutual funds"), None)
        stcg_mf_total = 0.0
        ltcg_mf_total = 0.0

        if mf_sheet:
            rows = list(mf_sheet.iter_rows(values_only=True))
            r = _find_row_by_label(rows, "Short Term profit Equity", col=1)
            if r:
                stcg_mf_total = _to_float(r[2] if len(r) > 2 else None)
            r = _find_row_by_label(rows, "Long Term profit Equity", col=1)
            if r:
                ltcg_mf_total = _to_float(r[2] if len(r) > 2 else None)

        # ── Per-trade breakdown from Tradewise sheet ──────────────────────────
        tradewise_name = next(
            (n for n in wb.sheetnames if "tradewise" in n.lower()), None
        )
        gains: list[dict] = []

        if tradewise_name:
            rows = list(wb[tradewise_name].iter_rows(values_only=True))
            # Header row has: Symbol, ISIN, Entry Date, Exit Date, Quantity,
            #                 Buy Value, Sell Value, Profit, Period of Holding, ...
            hdr_idx = _find_header_row(rows, ["symbol", "isin", "entry date", "exit date"], col_start=1)
            if hdr_idx is not None:
                # Build header map from col index
                headers = [str(c or "").strip().lower() for c in rows[hdr_idx]]

                def _col(name: str) -> Optional[int]:
                    for i, h in enumerate(headers):
                        if name in h:
                            return i
                    return None

                c_symbol  = _col("symbol")
                c_isin    = _col("isin")
                c_entry   = _col("entry date")
                c_exit    = _col("exit date")
                c_qty     = _col("quantity")
                c_buy_val = _col("buy value")
                c_sell_val= _col("sell value")
                c_profit  = _col("profit")
                c_days    = _col("period of holding")
                c_taxable = _col("taxable profit")

                for row in rows[hdr_idx + 1:]:
                    if not row or all(v is None for v in row):
                        continue
                    symbol   = str(row[c_symbol]  or "") if c_symbol  is not None else ""
                    isin     = str(row[c_isin]    or "") if c_isin    is not None else ""
                    profit   = _to_float(row[c_profit]   if c_profit   is not None else None)
                    taxable  = _to_float(row[c_taxable]  if c_taxable  is not None else profit)
                    buy_val  = _to_float(row[c_buy_val]  if c_buy_val  is not None else None)
                    sell_val = _to_float(row[c_sell_val] if c_sell_val is not None else None)
                    days     = _to_float(row[c_days]     if c_days     is not None else None)
                    entry    = str(row[c_entry] or "")   if c_entry    is not None else ""
                    exit_    = str(row[c_exit]  or "")   if c_exit     is not None else ""

                    if not symbol and not isin:
                        continue

                    asset_type = _isin_to_asset_type(isin, symbol)

                    if days == 0:
                        gain_type = GainType.STCG  # intraday — covered by speculation_total
                        is_speculation = True
                    elif days < 365:
                        gain_type = GainType.STCG
                        is_speculation = False
                    else:
                        gain_type = GainType.LTCG
                        is_speculation = False

                    gains.append({
                        "gain_type": gain_type,
                        "asset_type": asset_type,
                        "isin": isin,
                        "scrip": symbol,
                        "gain_amount": taxable,
                        "buy_date": entry,
                        "sell_date": exit_,
                        "buy_value": buy_val,
                        "sell_value": sell_val,
                        "holding_days": int(days),
                        "is_speculation": is_speculation,
                    })
            else:
                doc.parse_errors.append("Zerodha XLSX: could not find header row in Tradewise sheet")

        # ── Populate ParsedDocument ───────────────────────────────────────────
        self._set_summary_fields(
            doc, stcg_equity_total, ltcg_equity_total,
            speculation_total, fno_total, stcg_mf_total, ltcg_mf_total,
            gains, source_section="xlsx",
        )

    # ── PDF — legacy format ───────────────────────────────────────────────────

    def _parse_pdf(self, filepath: str, doc: ParsedDocument) -> None:
        try:
            pdf = pdfplumber.open(filepath)
        except Exception as exc:
            doc.parse_errors.append(f"Cannot open PDF: {exc}")
            return

        all_rows: list[list] = []
        with pdf:
            for page in pdf.pages:
                try:
                    for table in (page.extract_tables() or []):
                        all_rows.extend(table or [])
                except Exception as exc:
                    doc.parse_errors.append(f"Page table error: {exc}")

        self._process_table_rows(all_rows, doc, source_section="pdf")

    # ── CSV — legacy format ───────────────────────────────────────────────────

    def _parse_csv(self, filepath: str, doc: ParsedDocument) -> None:
        try:
            with open(filepath, "r", encoding="utf-8", errors="replace") as f:
                content = f.read()
        except Exception as exc:
            doc.parse_errors.append(f"Cannot read CSV: {exc}")
            return

        reader = csv.DictReader(io.StringIO(content))
        headers = reader.fieldnames or []
        rows = [list(headers)] + [[row.get(h, "") for h in headers] for row in reader]
        self._process_table_rows(rows, doc, source_section="csv")

    def _process_table_rows(self, rows: list, doc: ParsedDocument, source_section: str) -> None:
        """Shared handler for flat table formats (PDF / CSV)."""
        if not rows:
            doc.parse_errors.append("No rows found in Zerodha P&L")
            return

        _COL_ALIASES = {
            "scrip":     ["scrip", "symbol", "security"],
            "isin":      ["isin"],
            "buy_date":  ["buy date", "entry date", "purchase date"],
            "sell_date": ["sell date", "exit date", "sale date"],
            "buy_value": ["buy value", "purchase value"],
            "sell_value":["sell value", "sale value"],
            "stcg":      ["stcg", "short term", "st gain"],
            "ltcg":      ["ltcg", "long term", "lt gain"],
            "days":      ["period of holding", "days"],
        }

        header_idx = None
        headers: list[str] = []
        for i, row in enumerate(rows):
            row_lower = [str(c or "").strip().lower() for c in row]
            if "isin" in row_lower and any(
                k in row_lower for k in ["stcg", "ltcg", "short term", "long term"]
            ):
                header_idx = i
                headers = [str(c or "").strip() for c in row]
                break

        if header_idx is None:
            doc.parse_errors.append("Cannot find header row in Zerodha P&L table")
            return

        def _map(aliases):
            for h in headers:
                if h.strip().lower() in aliases:
                    return h
            return None

        col = {k: _map(v) for k, v in _COL_ALIASES.items()}

        gains: list[dict] = []
        stcg_total = 0.0
        ltcg_total = 0.0

        for row in rows[header_idx + 1:]:
            if not row or all(str(c or "").strip() == "" for c in row):
                continue
            rd = {h: str(row[i] if i < len(row) else "") for i, h in enumerate(headers)}

            isin = rd.get(col["isin"] or "", "")
            stcg_val = _to_float(rd.get(col["stcg"] or "", ""))
            ltcg_val = _to_float(rd.get(col["ltcg"] or "", ""))

            for gain_type, amount in [(GainType.STCG, stcg_val), (GainType.LTCG, ltcg_val)]:
                if amount != 0:
                    if gain_type == GainType.STCG:
                        stcg_total += amount
                    else:
                        ltcg_total += amount
                    gains.append({
                        "gain_type": gain_type,
                        "asset_type": _isin_to_asset_type(isin),
                        "isin": isin,
                        "scrip": rd.get(col["scrip"] or "", ""),
                        "gain_amount": amount,
                        "buy_date": rd.get(col["buy_date"] or "", ""),
                        "sell_date": rd.get(col["sell_date"] or "", ""),
                        "buy_value": _to_float(rd.get(col["buy_value"] or "", "")),
                        "sell_value": _to_float(rd.get(col["sell_value"] or "", "")),
                    })

        self._set_summary_fields(
            doc, stcg_total, ltcg_total, 0.0, 0.0, 0.0, 0.0,
            gains, source_section=source_section,
        )

    # ── Field population ──────────────────────────────────────────────────────

    def _set_summary_fields(
        self,
        doc: ParsedDocument,
        stcg_equity: float,
        ltcg_equity: float,
        speculation: float,
        fno: float,
        stcg_mf: float,
        ltcg_mf: float,
        gains: list[dict],
        source_section: str,
    ) -> None:
        conf = _EXACT_CONF

        doc.fields["stcg_equity_total"] = ExtractedField(
            value=stcg_equity, source_document=doc.filename,
            source_section=source_section, confidence=conf,
        )
        doc.fields["ltcg_equity_total"] = ExtractedField(
            value=ltcg_equity, source_document=doc.filename,
            source_section=source_section, confidence=conf,
        )
        doc.fields["speculation_total"] = ExtractedField(
            value=speculation, source_document=doc.filename,
            source_section=source_section, confidence=conf,
        )
        doc.fields["fno_total"] = ExtractedField(
            value=fno, source_document=doc.filename,
            source_section=source_section, confidence=conf,
        )
        doc.fields["stcg_mf_total"] = ExtractedField(
            value=stcg_mf, source_document=doc.filename,
            source_section=source_section, confidence=conf,
        )
        doc.fields["ltcg_mf_total"] = ExtractedField(
            value=ltcg_mf, source_document=doc.filename,
            source_section=source_section, confidence=conf,
        )
        doc.fields["capital_gains"] = ExtractedField(
            value=gains, source_document=doc.filename,
            source_section=source_section,
            confidence=conf if gains else _FUZZY_CONF,
        )

        if not gains and (stcg_equity == 0 and ltcg_equity == 0 and fno == 0):
            doc.parse_errors.append("No capital gains data extracted from Zerodha P&L")
